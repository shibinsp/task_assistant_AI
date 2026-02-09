"""
TaskPulse - AI Assistant - Report Service
Report generation and export functionality
"""

from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
import io
import json

from app.models.task import Task, TaskStatus, TaskPriority
from app.models.user import User
from app.models.checkin import CheckIn
from app.services.analytics_service import AnalyticsService
from app.utils.helpers import generate_uuid


class ReportService:
    """Service for generating reports."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.analytics = AnalyticsService(db)

    async def generate_team_productivity_report(
        self,
        org_id: str,
        team_id: str,
        start_date: datetime,
        end_date: datetime,
        format: str = "json"
    ) -> Dict[str, Any]:
        """Generate team productivity report."""
        # Get team members
        result = await self.db.execute(
            select(User).where(
                User.org_id == org_id,
                User.team_id == team_id,
                User.is_active == True
            )
        )
        team_members = result.scalars().all()

        # Get tasks in period
        result = await self.db.execute(
            select(Task).where(
                Task.org_id == org_id,
                Task.team_id == team_id,
                Task.created_at >= start_date,
                Task.created_at <= end_date
            )
        )
        tasks = result.scalars().all()

        # Calculate metrics per member
        member_metrics = []
        for member in team_members:
            member_tasks = [t for t in tasks if t.assigned_to == member.id]
            completed = [t for t in member_tasks if t.is_completed]
            on_time = [t for t in completed if t.deadline and t.completed_at and t.completed_at <= t.deadline]

            member_metrics.append({
                "user_id": member.id,
                "name": f"{member.first_name} {member.last_name}",
                "tasks_assigned": len(member_tasks),
                "tasks_completed": len(completed),
                "completion_rate": round(len(completed) / max(1, len(member_tasks)) * 100, 1),
                "on_time_rate": round(len(on_time) / max(1, len(completed)) * 100, 1),
                "total_hours": sum(t.actual_hours or 0 for t in completed)
            })

        # Team summary
        total_tasks = len(tasks)
        total_completed = sum(m["tasks_completed"] for m in member_metrics)

        report_data = {
            "report_type": "team_productivity",
            "team_id": team_id,
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            },
            "generated_at": datetime.utcnow().isoformat(),
            "summary": {
                "team_size": len(team_members),
                "total_tasks": total_tasks,
                "total_completed": total_completed,
                "team_completion_rate": round(total_completed / max(1, total_tasks) * 100, 1),
                "avg_tasks_per_member": round(total_tasks / max(1, len(team_members)), 1)
            },
            "member_breakdown": member_metrics
        }

        if format == "json":
            return report_data
        elif format == "csv":
            return self._to_csv(report_data, "member_breakdown")
        elif format == "xlsx":
            return await self._to_xlsx(report_data)

        return report_data

    async def generate_task_completion_report(
        self,
        org_id: str,
        start_date: datetime,
        end_date: datetime,
        team_id: Optional[str] = None,
        format: str = "json"
    ) -> Dict[str, Any]:
        """Generate task completion analysis report."""
        filters = [
            Task.org_id == org_id,
            Task.completed_at >= start_date,
            Task.completed_at <= end_date,
            Task.status == TaskStatus.DONE
        ]
        if team_id:
            filters.append(Task.team_id == team_id)

        result = await self.db.execute(select(Task).where(*filters))
        completed_tasks = result.scalars().all()

        # Group by priority
        by_priority = {}
        for priority in TaskPriority:
            priority_tasks = [t for t in completed_tasks if t.priority == priority]
            if priority_tasks:
                avg_time = sum((t.completed_at - t.started_at).total_seconds() / 3600
                              for t in priority_tasks if t.started_at and t.completed_at) / len(priority_tasks)
                by_priority[priority.value] = {
                    "count": len(priority_tasks),
                    "avg_completion_hours": round(avg_time, 1)
                }

        # Time analysis
        on_time = [t for t in completed_tasks if t.deadline and t.completed_at <= t.deadline]
        early = [t for t in on_time if t.deadline and (t.deadline - t.completed_at).days > 1]
        late = [t for t in completed_tasks if t.deadline and t.completed_at > t.deadline]

        report_data = {
            "report_type": "task_completion",
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            },
            "generated_at": datetime.utcnow().isoformat(),
            "summary": {
                "total_completed": len(completed_tasks),
                "on_time": len(on_time),
                "early": len(early),
                "late": len(late),
                "on_time_rate": round(len(on_time) / max(1, len([t for t in completed_tasks if t.deadline])) * 100, 1)
            },
            "by_priority": by_priority,
            "tasks": [
                {
                    "id": t.id,
                    "title": t.title,
                    "priority": t.priority.value,
                    "completed_at": t.completed_at.isoformat() if t.completed_at else None,
                    "deadline": t.deadline.isoformat() if t.deadline else None,
                    "was_on_time": t.deadline and t.completed_at and t.completed_at <= t.deadline if t.deadline else None
                }
                for t in completed_tasks[:100]  # Limit to 100 for readability
            ]
        }

        return report_data

    async def generate_blocker_analysis_report(
        self,
        org_id: str,
        start_date: datetime,
        end_date: datetime,
        format: str = "json"
    ) -> Dict[str, Any]:
        """Generate blocker analysis report."""
        # Get check-ins with blockers
        result = await self.db.execute(
            select(CheckIn).where(
                CheckIn.org_id == org_id,
                CheckIn.created_at >= start_date,
                CheckIn.created_at <= end_date,
                CheckIn.has_blocker == True
            )
        )
        blocked_checkins = result.scalars().all()

        # Group by blocker type
        by_type = {}
        for checkin in blocked_checkins:
            blocker_type = checkin.blocker_type.value if checkin.blocker_type else "unknown"
            if blocker_type not in by_type:
                by_type[blocker_type] = {
                    "count": 0,
                    "resolved": 0,
                    "avg_resolution_hours": 0,
                    "examples": []
                }
            by_type[blocker_type]["count"] += 1
            if checkin.resolved_at:
                by_type[blocker_type]["resolved"] += 1
            if len(by_type[blocker_type]["examples"]) < 3:
                by_type[blocker_type]["examples"].append(checkin.blocker_description or "No description")

        # Get currently blocked tasks
        result = await self.db.execute(
            select(Task).where(
                Task.org_id == org_id,
                Task.status == TaskStatus.BLOCKED
            )
        )
        currently_blocked = result.scalars().all()

        report_data = {
            "report_type": "blocker_analysis",
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            },
            "generated_at": datetime.utcnow().isoformat(),
            "summary": {
                "total_blockers_reported": len(blocked_checkins),
                "currently_blocked_tasks": len(currently_blocked),
                "unique_blocker_types": len(by_type)
            },
            "by_type": by_type,
            "currently_blocked": [
                {
                    "id": t.id,
                    "title": t.title,
                    "blocker_type": t.blocker_type.value if t.blocker_type else None,
                    "description": t.blocker_description,
                    "blocked_since": t.updated_at.isoformat()
                }
                for t in currently_blocked
            ],
            "recommendations": self._generate_blocker_recommendations(by_type)
        }

        return report_data

    def _generate_blocker_recommendations(self, by_type: Dict[str, Any]) -> List[str]:
        """Generate recommendations based on blocker analysis."""
        recommendations = []

        if by_type.get("dependency", {}).get("count", 0) > 5:
            recommendations.append("High dependency blockers - consider better task sequencing")

        if by_type.get("resource", {}).get("count", 0) > 3:
            recommendations.append("Resource constraints detected - review resource allocation")

        if by_type.get("tool", {}).get("count", 0) > 3:
            recommendations.append("Tool-related blockers - evaluate tooling improvements")

        if not recommendations:
            recommendations.append("Blocker levels are within normal range")

        return recommendations

    async def generate_executive_summary(
        self,
        org_id: str,
        period_days: int = 30
    ) -> Dict[str, Any]:
        """Generate executive summary report."""
        end_date = datetime.utcnow()
        start_date = end_date - timedelta(days=period_days)

        # Get dashboard metrics
        dashboard = await self.analytics.get_dashboard_metrics(org_id)

        # Get velocity data
        velocity = await self.analytics.get_velocity_chart_data(org_id, weeks=8)

        # Get bottleneck analysis
        bottlenecks = await self.analytics.get_bottleneck_analysis(org_id)

        # Employee count
        result = await self.db.execute(
            select(func.count()).where(User.org_id == org_id, User.is_active == True)
        )
        employee_count = result.scalar() or 0

        return {
            "report_type": "executive_summary",
            "organization_id": org_id,
            "period_days": period_days,
            "generated_at": datetime.utcnow().isoformat(),
            "key_metrics": {
                "total_employees": employee_count,
                "active_tasks": dashboard["task_summary"]["total_active"],
                "completed_this_month": dashboard["completion_metrics"]["completed_this_month"],
                "on_time_rate": dashboard["completion_metrics"]["on_time_rate"],
                "blocked_tasks": dashboard["blocker_analysis"]["total_blocked"],
                "health_score": bottlenecks["health_score"]
            },
            "velocity_trend": velocity["trend"],
            "velocity_average": velocity["average"],
            "top_bottlenecks": [
                {"type": b["type"], "severity": b["severity"], "count": b["count"]}
                for b in bottlenecks["bottlenecks"][:3]
            ],
            "executive_insights": self._generate_executive_insights(dashboard, velocity, bottlenecks)
        }

    def _generate_executive_insights(
        self,
        dashboard: Dict,
        velocity: Dict,
        bottlenecks: Dict
    ) -> List[str]:
        """Generate executive-level insights."""
        insights = []

        # Completion rate insight
        on_time_rate = dashboard["completion_metrics"]["on_time_rate"]
        if on_time_rate >= 90:
            insights.append(f"Excellent on-time delivery rate at {on_time_rate}%")
        elif on_time_rate >= 75:
            insights.append(f"Good on-time delivery rate at {on_time_rate}%, room for improvement")
        else:
            insights.append(f"On-time delivery rate at {on_time_rate}% needs attention")

        # Velocity trend
        if velocity["trend"] == "increasing":
            insights.append("Team velocity is trending upward")
        elif velocity["trend"] == "decreasing":
            insights.append("Team velocity is declining - investigate root causes")

        # Bottleneck insight
        if bottlenecks["bottleneck_count"] == 0:
            insights.append("No significant bottlenecks detected")
        else:
            insights.append(f"{bottlenecks['bottleneck_count']} bottleneck(s) require attention")

        return insights

    def _to_csv(self, data: Dict[str, Any], list_key: str) -> str:
        """Convert report data to CSV format."""
        if list_key not in data or not data[list_key]:
            return ""

        items = data[list_key]
        if not items:
            return ""

        # Get headers from first item
        headers = list(items[0].keys())
        lines = [",".join(headers)]

        for item in items:
            values = [str(item.get(h, "")) for h in headers]
            lines.append(",".join(values))

        return "\n".join(lines)

    async def _to_xlsx(self, data: Dict[str, Any]) -> bytes:
        """Convert report data to XLSX format."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            # Add title
            ws['A1'] = data.get("report_type", "Report")
            ws['A1'].font = Font(bold=True, size=14)

            # Add summary
            row = 3
            ws[f'A{row}'] = "Summary"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            summary = data.get("summary", {})
            for key, value in summary.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1

            # Add member breakdown if exists
            if "member_breakdown" in data:
                row += 2
                ws[f'A{row}'] = "Member Breakdown"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

                members = data["member_breakdown"]
                if members:
                    headers = list(members[0].keys())
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=header).font = Font(bold=True)
                    row += 1

                    for member in members:
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=member.get(header))
                        row += 1

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()

        except ImportError:
            # If openpyxl not available, return JSON
            return json.dumps(data).encode()

    async def schedule_report(
        self,
        org_id: str,
        report_type: str,
        schedule: str,  # daily, weekly, monthly
        recipients: List[str],
        config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Schedule a recurring report (stub for background job integration)."""
        # This would integrate with APScheduler or similar
        return {
            "scheduled": True,
            "report_type": report_type,
            "schedule": schedule,
            "recipients": recipients,
            "next_run": (datetime.utcnow() + timedelta(days=1)).isoformat(),
            "message": "Report scheduled successfully"
        }
